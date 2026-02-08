"""
Multi-sheet Excel report generator for budget analysis.
"""

from typing import List, Dict, Any, Optional
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference, PieChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelReportGenerator:
    """
    Generates multi-sheet Excel reports from budget data.

    Sheets:
    1. Summary - High-level statistics
    2. All Items - Full data table
    3. By Agency - Aggregated by agency
    4. By Category - Aggregated by program category
    5. By Year - Aggregated by chapter year
    6. Charts Data - Data formatted for charting
    """

    def __init__(self, records: List[Dict[str, Any]]):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        self.records = records
        self.wb = Workbook()

        # Styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.money_format = '#,##0'
        self.pct_format = '0.0%'
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(self, output_path: str, title: str = "NYS Budget Analysis"):
        """Generate complete Excel report."""
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets
        self._create_summary_sheet(title)
        self._create_data_sheet()
        self._create_agency_sheet()
        self._create_category_sheet()
        self._create_year_sheet()
        self._create_charts_data_sheet()

        # Save
        self.wb.save(output_path)
        return output_path

    def _create_summary_sheet(self, title: str):
        """Create summary statistics sheet."""
        ws = self.wb.create_sheet("Summary", 0)

        # Calculate stats
        total_records = len(self.records)
        total_amount = sum(self._get_amount(r) for r in self.records)
        unique_agencies = len(set(r.get('agency', '') for r in self.records))
        unique_ids = len(set(r.get('appropriation_id', '') for r in self.records))

        # Get year range
        years = [r.get('chapter_year') for r in self.records if r.get('chapter_year')]
        years = [y for y in years if y and str(y).isdigit()]
        min_year = min(years) if years else 'N/A'
        max_year = max(years) if years else 'N/A'

        # Title
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Summary stats
        stats = [
            ('Total Records', total_records),
            ('Total Amount', total_amount),
            ('Unique Agencies', unique_agencies),
            ('Unique Appropriation IDs', unique_ids),
            ('Chapter Year Range', f"{min_year} - {max_year}"),
        ]

        ws['A3'] = "Summary Statistics"
        ws['A3'].font = Font(bold=True, size=12)

        for i, (label, value) in enumerate(stats, start=4):
            ws[f'A{i}'] = label
            ws[f'B{i}'] = value
            if isinstance(value, (int, float)) and 'Amount' in label:
                ws[f'B{i}'].number_format = self.money_format

        # Top 10 agencies
        agency_totals = self._aggregate_by_field('agency')
        top_agencies = sorted(agency_totals.items(), key=lambda x: x[1]['total'], reverse=True)[:10]

        ws['A10'] = "Top 10 Agencies by Amount"
        ws['A10'].font = Font(bold=True, size=12)

        headers = ['Rank', 'Agency', 'Count', 'Total Amount', '% of Total']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for i, (agency, data) in enumerate(top_agencies, start=1):
            row = 11 + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=agency[:50])  # Truncate long names
            ws.cell(row=row, column=3, value=data['count'])
            ws.cell(row=row, column=4, value=data['total']).number_format = self.money_format
            ws.cell(row=row, column=5, value=data['total'] / total_amount if total_amount else 0).number_format = self.pct_format

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12

    def _create_data_sheet(self):
        """Create sheet with all data."""
        ws = self.wb.create_sheet("All Items")

        if not self.records:
            ws['A1'] = "No data"
            return

        # Determine columns to include
        priority_cols = [
            'agency', 'appropriation_id', 'chapter_year',
            'appropriation_amount', 'reappropriation_amount',
            'budget_type', 'fund_type', 'account',
            'program_category', 'recipient_name',
            'page_number'
        ]

        # Get all available columns
        all_cols = set()
        for r in self.records:
            all_cols.update(r.keys())

        # Order: priority cols first, then others
        columns = [c for c in priority_cols if c in all_cols]
        columns.extend(sorted(c for c in all_cols if c not in priority_cols and c != 'bill_language'))

        # Write headers
        for col, header in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Write data
        for row_idx, record in enumerate(self.records, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = record.get(col_name, '')

                # Handle lists/dicts
                if isinstance(value, (list, dict)):
                    value = str(value)

                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format amounts
                if 'amount' in col_name.lower() and isinstance(value, (int, float)):
                    cell.number_format = self.money_format

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Auto-filter
        ws.auto_filter.ref = f"A1:{chr(65 + len(columns) - 1)}{len(self.records) + 1}"

    def _create_agency_sheet(self):
        """Create aggregated by-agency sheet."""
        ws = self.wb.create_sheet("By Agency")

        agency_data = self._aggregate_by_field('agency')
        sorted_data = sorted(agency_data.items(), key=lambda x: x[1]['total'], reverse=True)

        headers = ['Agency', 'Item Count', 'Total Amount', 'Average Amount', '% of Total']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        total_amount = sum(d['total'] for _, d in sorted_data)

        for row_idx, (agency, data) in enumerate(sorted_data, start=2):
            ws.cell(row=row_idx, column=1, value=agency)
            ws.cell(row=row_idx, column=2, value=data['count'])
            ws.cell(row=row_idx, column=3, value=data['total']).number_format = self.money_format
            ws.cell(row=row_idx, column=4, value=data['avg']).number_format = self.money_format
            pct = data['total'] / total_amount if total_amount else 0
            ws.cell(row=row_idx, column=5, value=pct).number_format = self.pct_format

        # Column widths
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f"A1:E{len(sorted_data) + 1}"

    def _create_category_sheet(self):
        """Create aggregated by-category sheet."""
        ws = self.wb.create_sheet("By Category")

        cat_data = self._aggregate_by_field('program_category')
        sorted_data = sorted(cat_data.items(), key=lambda x: x[1]['total'], reverse=True)

        headers = ['Category', 'Item Count', 'Total Amount', 'Average Amount', '% of Total']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        total_amount = sum(d['total'] for _, d in sorted_data)

        for row_idx, (category, data) in enumerate(sorted_data, start=2):
            cat_display = category if category else '(Uncategorized)'
            ws.cell(row=row_idx, column=1, value=cat_display)
            ws.cell(row=row_idx, column=2, value=data['count'])
            ws.cell(row=row_idx, column=3, value=data['total']).number_format = self.money_format
            ws.cell(row=row_idx, column=4, value=data['avg']).number_format = self.money_format
            pct = data['total'] / total_amount if total_amount else 0
            ws.cell(row=row_idx, column=5, value=pct).number_format = self.pct_format

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12

    def _create_year_sheet(self):
        """Create aggregated by-year sheet."""
        ws = self.wb.create_sheet("By Year")

        year_data = self._aggregate_by_field('chapter_year')
        # Sort by year descending
        sorted_data = sorted(year_data.items(), key=lambda x: str(x[0]), reverse=True)

        headers = ['Chapter Year', 'Item Count', 'Total Amount', 'Average Amount']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill

        for row_idx, (year, data) in enumerate(sorted_data, start=2):
            ws.cell(row=row_idx, column=1, value=year if year else '(No Year)')
            ws.cell(row=row_idx, column=2, value=data['count'])
            ws.cell(row=row_idx, column=3, value=data['total']).number_format = self.money_format
            ws.cell(row=row_idx, column=4, value=data['avg']).number_format = self.money_format

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    def _create_charts_data_sheet(self):
        """Create sheet with data formatted for charts."""
        ws = self.wb.create_sheet("Charts Data")

        # Top 10 agencies for bar chart
        ws['A1'] = "Top 10 Agencies"
        ws['A1'].font = Font(bold=True)

        agency_data = self._aggregate_by_field('agency')
        top_agencies = sorted(agency_data.items(), key=lambda x: x[1]['total'], reverse=True)[:10]

        ws['A2'] = "Agency"
        ws['B2'] = "Amount (Millions)"
        ws['A2'].font = self.header_font
        ws['A2'].fill = self.header_fill
        ws['B2'].font = self.header_font
        ws['B2'].fill = self.header_fill

        for i, (agency, data) in enumerate(top_agencies, start=3):
            ws.cell(row=i, column=1, value=agency[:40])
            ws.cell(row=i, column=2, value=data['total'] / 1_000_000)

        # Category breakdown for pie chart
        ws['D1'] = "By Category"
        ws['D1'].font = Font(bold=True)

        cat_data = self._aggregate_by_field('program_category')
        sorted_cats = sorted(cat_data.items(), key=lambda x: x[1]['total'], reverse=True)

        ws['D2'] = "Category"
        ws['E2'] = "Amount"
        ws['D2'].font = self.header_font
        ws['D2'].fill = self.header_fill
        ws['E2'].font = self.header_font
        ws['E2'].fill = self.header_fill

        for i, (cat, data) in enumerate(sorted_cats, start=3):
            cat_display = cat if cat else '(Uncategorized)'
            ws.cell(row=i, column=4, value=cat_display)
            ws.cell(row=i, column=5, value=data['total'])

        # Year trend
        ws['G1'] = "By Year"
        ws['G1'].font = Font(bold=True)

        year_data = self._aggregate_by_field('chapter_year')
        sorted_years = sorted(
            [(y, d) for y, d in year_data.items() if y and str(y).isdigit()],
            key=lambda x: str(x[0])
        )

        ws['G2'] = "Year"
        ws['H2'] = "Amount"
        ws['G2'].font = self.header_font
        ws['G2'].fill = self.header_fill
        ws['H2'].font = self.header_font
        ws['H2'].fill = self.header_fill

        for i, (year, data) in enumerate(sorted_years, start=3):
            ws.cell(row=i, column=7, value=int(year))
            ws.cell(row=i, column=8, value=data['total'])

        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 18

    def _aggregate_by_field(self, field: str) -> Dict[str, Dict[str, Any]]:
        """Aggregate records by a field."""
        groups = defaultdict(list)

        for r in self.records:
            key = r.get(field, '')
            groups[key].append(self._get_amount(r))

        result = {}
        for key, amounts in groups.items():
            result[key] = {
                'count': len(amounts),
                'total': sum(amounts),
                'avg': sum(amounts) / len(amounts) if amounts else 0,
                'min': min(amounts) if amounts else 0,
                'max': max(amounts) if amounts else 0,
            }

        return result

    def _get_amount(self, record: Dict[str, Any]) -> float:
        """Extract amount from record."""
        for field in ['reappropriation_amount', 'appropriation_amount']:
            val = record.get(field)
            if val:
                try:
                    return float(val)
                except (ValueError, TypeError):
                    pass
        return 0


def generate_excel_report(records: List[Dict[str, Any]],
                          output_path: str,
                          title: str = "NYS Budget Analysis") -> str:
    """
    Generate an Excel report from budget records.

    Args:
        records: List of budget record dictionaries
        output_path: Path for output Excel file
        title: Report title

    Returns:
        Path to generated file
    """
    generator = ExcelReportGenerator(records)
    return generator.generate(output_path, title)
